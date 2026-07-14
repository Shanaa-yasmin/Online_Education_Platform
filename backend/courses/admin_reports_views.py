import csv
import hashlib
import io
import datetime
from django.core.cache import cache
from django.db.models.functions import TruncMonth, TruncDay
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.contrib.auth import get_user_model
from django.db.models import Sum, Avg, Count, Q, DecimalField
from django.db.models.functions import Coalesce, TruncMonth, TruncDay
from django.utils import timezone

from courses.models import Course, Lesson
from payments.models import Enrollment, Payment
from certificates.models import Certificate
from notifications.models import Notification

User = get_user_model()

# ── Cache constants ────────────────────────────────────────────────────────────
# ADMIN_REPORTS_BASE_KEY is the stable cache key for the unfiltered reports
# response.  Invalidation callers (course approve/reject) delete this key so
# admins see fresh stats immediately.  Filtered variants (with query params)
# use per-fingerprint keys and expire naturally at ADMIN_REPORTS_CACHE_TTL.
ADMIN_REPORTS_BASE_KEY = "admin_reports:nofilter"
ADMIN_REPORTS_CACHE_TTL = 300  # 5 minutes — reports are analytical/historical;
                                # users don't expect real-time freshness here.


def _make_reports_cache_key(request) -> str:
    """Return a stable cache key for the current set of query-param filters.

    When there are no filters, returns ADMIN_REPORTS_BASE_KEY so invalidation
    callers have a predictable key to delete.  When filters are present, builds
    a short MD5 fingerprint of the sorted params so each unique combination gets
    its own cache slot without creating unbounded key sprawl.
    """
    params = dict(sorted(request.query_params.items()))
    if not params:
        return ADMIN_REPORTS_BASE_KEY
    fingerprint = hashlib.md5(str(params).encode(), usedforsecurity=False).hexdigest()[:12]
    return f"admin_reports:{fingerprint}"

try:
    import openpyxl
    from openpyxl.styles import Font as xlFont, PatternFill as xlPatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def get_admin_reports_data(request):
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    course_id = request.query_params.get('course_id')
    mentor_id = request.query_params.get('mentor_id')
    student_id = request.query_params.get('student_id')
    category = request.query_params.get('category')
    min_revenue = request.query_params.get('min_revenue')
    max_revenue = request.query_params.get('max_revenue')
    completion_status = request.query_params.get('completion_status')

    # Base querysets
    enrollments = Enrollment.objects.all().select_related('student', 'course', 'course__mentor')
    payments = Payment.objects.all().select_related('student', 'enrollment__course')

    # Apply filters
    if start_date:
        enrollments = enrollments.filter(enrolled_at__date__gte=start_date)
        payments = payments.filter(created_at__date__gte=start_date)
    if end_date:
        enrollments = enrollments.filter(enrolled_at__date__lte=end_date)
        payments = payments.filter(created_at__date__lte=end_date)
    if course_id:
        enrollments = enrollments.filter(course_id=course_id)
        payments = payments.filter(enrollment__course_id=course_id)
    if mentor_id:
        enrollments = enrollments.filter(course__mentor_id=mentor_id)
        payments = payments.filter(enrollment__course__mentor_id=mentor_id)
    if student_id:
        enrollments = enrollments.filter(student_id=student_id)
        payments = payments.filter(student_id=student_id)
    if category:
        enrollments = enrollments.filter(course__category__iexact=category)
        payments = payments.filter(enrollment__course__category__iexact=category)
    if min_revenue:
        payments = payments.filter(amount__gte=min_revenue)
    if max_revenue:
        payments = payments.filter(amount__lte=max_revenue)
    if completion_status == 'completed':
        enrollments = enrollments.filter(progress_percent__gte=100.0)
    elif completion_status == 'in_progress':
        enrollments = enrollments.filter(progress_percent__lt=100.0)

    # Aggregations
    user_stats = User.objects.aggregate(
        total=Count('id'),
        students=Count('id', filter=Q(role='STUDENT')),
        mentors=Count('id', filter=Q(role='MENTOR')),
        admins=Count('id', filter=Q(role='ADMIN') | Q(is_staff=True))
    )
    total_users = user_stats['total']
    students_count = user_stats['students']
    mentors_count = user_stats['mentors']
    admins_count = user_stats['admins']

    course_stats = Course.objects.aggregate(
        total=Count('id'),
        published=Count('id', filter=Q(is_published=True)),
        pending=Count('id', filter=Q(is_submitted_for_review=True, is_approved=False)),
        rejected=Count('id', filter=Q(is_rejected=True))
    )
    total_courses = course_stats['total']
    published_courses = course_stats['published']
    pending_courses = course_stats['pending']
    rejected_courses = course_stats['rejected']

    total_enrollments = enrollments.count()

    # Combine revenue + refunds into a single grouped query
    payment_sums = payments.filter(
        status__in=[Payment.StatusChoices.COMPLETED, Payment.StatusChoices.REFUNDED]
    ).values('status').annotate(total=Sum('amount'))
    payment_sums_dict = {item['status']: float(item['total'] or 0.0) for item in payment_sums}
    revenue = payment_sums_dict.get(Payment.StatusChoices.COMPLETED, 0.0)
    refunds = payment_sums_dict.get(Payment.StatusChoices.REFUNDED, 0.0)

    avg_rating = Course.objects.aggregate(avg=Avg('rating_average'))['avg'] or 0.0
    avg_rating = round(float(avg_rating), 2)

    certificates_issued = Certificate.objects.filter(enrollment__in=enrollments).count()
    notifications_sent = Notification.objects.count()

    daily_active_users = User.objects.filter(
        last_login__gte=timezone.now() - datetime.timedelta(days=1)
    ).count()

    # List data
    # Monthly enrollments
    monthly_data = list(
        Enrollment.objects.annotate(month=TruncMonth('enrolled_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')[:12]
    )
    monthly_enrollments = []
    for md in monthly_data:
        monthly_enrollments.append({
            "month": md['month'].strftime('%Y-%b') if md['month'] else '',
            "enrollments": md['count']
        })

    # Top courses — select_related('mentor') avoids N+1 user lookups
    top_courses_qs = Course.objects.select_related('mentor').annotate(
        active_students=Count('enrollments', filter=Q(enrollments__is_active=True))
    ).order_by('-active_students')[:5]
    top_courses = []
    for tc in top_courses_qs:
        top_courses.append({
            "id": tc.id,
            "title": tc.title,
            "students": tc.active_students,
            "mentor": tc.mentor.username
        })

    # Top Mentors — single annotated query replaces in-Python sort of all mentors
    top_mentors_qs = User.objects.filter(role='MENTOR').annotate(
        revenue=Coalesce(
            Sum(
                'created_courses__enrollments__payments__amount',
                filter=Q(created_courses__enrollments__payments__status=Payment.StatusChoices.COMPLETED)
            ),
            0.0,
            output_field=DecimalField()
        )
    ).order_by('-revenue')[:5]
    top_mentors = [
        {"id": tm.id, "username": tm.username, "revenue": float(tm.revenue)}
        for tm in top_mentors_qs
    ]

    # Most Active Students
    most_active_qs = User.objects.filter(role='STUDENT').annotate(
        completed_count=Count('lesson_progresses', filter=Q(lesson_progresses__completed=True))
    ).order_by('-completed_count')[:5]
    most_active_students = []
    for ma in most_active_qs:
        most_active_students.append({
            "id": ma.id,
            "username": ma.username,
            "completed_lessons": ma.completed_count
        })

    # Most Popular Categories
    category_counts = list(
        Course.objects.values('category')
        .annotate(students=Count('enrollments', filter=Q(enrollments__is_active=True)))
        .order_by('-students')[:5]
    )

    stats = {
        "total_users": total_users,
        "students": students_count,
        "mentors": mentors_count,
        "admins": admins_count,
        "total_courses": total_courses,
        "published_courses": published_courses,
        "pending_courses": pending_courses,
        "rejected_courses": rejected_courses,
        "total_enrollments": total_enrollments,
        "revenue": revenue,
        "refunds": refunds,
        "avg_rating": avg_rating,
        "certificates_issued": certificates_issued,
        "notifications_sent": notifications_sent,
        "daily_active_users": daily_active_users,
    }

    return {
        "stats": stats,
        "monthly_enrollments": monthly_enrollments,
        "top_courses": top_courses,
        "top_mentors": top_mentors,
        "most_active_students": most_active_students,
        "most_popular_categories": category_counts,
        "enrollments": enrollments,
        "payments": payments
    }


class AdminReportsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'ADMIN':
            return Response({"detail": "Only administrators can view admin reports."}, status=status.HTTP_403_FORBIDDEN)

        cache_key = _make_reports_cache_key(request)
        cached_response = cache.get(cache_key)
        if cached_response is not None:
            return Response(cached_response, status=status.HTTP_200_OK)

        data = get_admin_reports_data(request)

        response_data = {
            "stats": data["stats"],
            "monthly_enrollments": data["monthly_enrollments"],
            "top_courses": data["top_courses"],
            "top_mentors": data["top_mentors"],
            "most_active_students": data["most_active_students"],
            "most_popular_categories": data["most_popular_categories"]
        }
        cache.set(cache_key, response_data, ADMIN_REPORTS_CACHE_TTL)
        return Response(response_data, status=status.HTTP_200_OK)


class AdminReportFilterOptionsView(APIView):
    """
    Provides lightweight lookup lists (id + display name) so the frontend
    can render dropdowns instead of asking admins to type raw IDs.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'ADMIN':
            return Response({"detail": "Only administrators can view admin reports."}, status=status.HTTP_403_FORBIDDEN)

        courses = list(Course.objects.order_by('title').values('id', 'title'))
        mentors = list(User.objects.filter(role='MENTOR').order_by('username').values('id', 'username'))
        students = list(User.objects.filter(role='STUDENT').order_by('username').values('id', 'username'))
        categories = list(
            Course.objects.exclude(category__isnull=True)
            .exclude(category__exact='')
            .order_by('category')
            .values_list('category', flat=True)
            .distinct()
        )

        return Response({
            "courses": courses,
            "mentors": mentors,
            "students": students,
            "categories": categories,
        }, status=status.HTTP_200_OK)


class AdminReportsExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != 'ADMIN':
            return HttpResponse("Access Denied", status=403)

        format_type = request.query_params.get('export_format', 'csv').lower()
        data = get_admin_reports_data(request)
        enrollments = data["enrollments"]
        stats = data["stats"]

        # Pre-fetch matching completed payments to avoid N+1 query loops
        payments_by_enrollment = {
            p.enrollment_id: p
            for p in Payment.objects.filter(
                enrollment__in=enrollments, status=Payment.StatusChoices.COMPLETED
            )
        }

        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="admin_report_{datetime.datetime.now().strftime("%Y%m%d")}.csv"'

            writer = csv.writer(response)
            writer.writerow(['Date Enrolled', 'Student Username', 'Student Email', 'Course Title', 'Category', 'Progress %', 'Revenue Amount'])

            for enroll in enrollments:
                pay = payments_by_enrollment.get(enroll.id)
                amt = float(pay.amount) if pay else 0.00
                writer.writerow([
                    enroll.enrolled_at.strftime('%Y-%m-%d %H:%M:%S'),
                    enroll.student.username,
                    enroll.student.email,
                    enroll.course.title,
                    enroll.course.category,
                    f"{enroll.progress_percent}%",
                    f"${amt:.2f}"
                ])
            return response

        elif format_type == 'excel' or format_type == 'xlsx':
            if HAS_OPENPYXL:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Administration Report"

                header_font = xlFont(name="Arial", size=11, bold=True, color="FFFFFF")
                header_fill = xlPatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

                headers = ['Date Enrolled', 'Student Username', 'Student Email', 'Course Title', 'Category', 'Progress %', 'Revenue Amount']
                ws.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill

                for enroll in enrollments:
                    pay = payments_by_enrollment.get(enroll.id)
                    amt = float(pay.amount) if pay else 0.00
                    ws.append([
                        enroll.enrolled_at.strftime('%Y-%m-%d'),
                        enroll.student.username,
                        enroll.student.email,
                        enroll.course.title,
                        enroll.course.category,
                        f"{enroll.progress_percent}%",
                        amt
                    ])

                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="admin_report_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx"'
                wb.save(response)
                return response
            else:
                response = HttpResponse(content_type='application/vnd.ms-excel')
                response['Content-Disposition'] = f'attachment; filename="admin_report_{datetime.datetime.now().strftime("%Y%m%d")}.xls"'

                writer = csv.writer(response, delimiter='\t')
                writer.writerow(['Date Enrolled', 'Student Username', 'Student Email', 'Course Title', 'Category', 'Progress %', 'Revenue Amount'])
                for enroll in enrollments:
                    pay = payments_by_enrollment.get(enroll.id)
                    amt = float(pay.amount) if pay else 0.00
                    writer.writerow([
                        enroll.enrolled_at.strftime('%Y-%m-%d'),
                        enroll.student.username,
                        enroll.student.email,
                        enroll.course.title,
                        enroll.course.category,
                        f"{enroll.progress_percent}%",
                        f"${amt:.2f}"
                    ])
                return response

        elif format_type == 'pdf':
            pdf_bytes = generate_pdf_report(enrollments, stats, payments_by_enrollment)
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="admin_report_{datetime.datetime.now().strftime("%Y%m%d")}.pdf"'
            return response

        return HttpResponse("Invalid format type.", status=400)


def generate_pdf_report(enrollments, stats, payments_by_enrollment=None):
    import os
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor("#1A365D"),
        spaceAfter=20
    )
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor("#2B6CB0"),
        spaceBefore=15,
        spaceAfter=10
    )
    body_style = styles['BodyText']

    # Logo (top of report)
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'report_assets', 'logo.jpeg')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=0.6*inch, height=0.6*inch)
        logo.hAlign = 'LEFT'
        story.append(logo)
        story.append(Spacer(1, 8))

    story.append(Paragraph("EduPath Platform Administration Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", body_style))
    story.append(Spacer(1, 15))

    story.append(Paragraph("Platform Overview Stats", section_style))
    stats_data = [
        ["Stat Category", "Value", "Stat Category", "Value"],
        ["Total Users", str(stats.get("total_users", 0)), "Total Courses", str(stats.get("total_courses", 0))],
        ["Students", str(stats.get("students", 0)), "Published Courses", str(stats.get("published_courses", 0))],
        ["Mentors", str(stats.get("mentors", 0)), "Pending Courses", str(stats.get("pending_courses", 0))],
        ["Total Enrollments", str(stats.get("total_enrollments", 0)), "Total Revenue", f"${stats.get('revenue', 0.0):.2f}"],
        ["Certificates Issued", str(stats.get("certificates_issued", 0)), "Total Refunds", f"${stats.get('refunds', 0.0):.2f}"]
    ]
    t_stats = Table(stats_data, colWidths=[130, 130, 130, 130])
    t_stats.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EDF2F7")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#2D3748")),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    story.append(t_stats)
    story.append(Spacer(1, 20))

    story.append(Paragraph(f"Filtered Transactions (Top 30 shown)", section_style))
    tx_data = [["Date", "Student", "Course", "Progress", "Amount"]]
    for enroll in enrollments[:30]:
        if payments_by_enrollment is not None:
            pay = payments_by_enrollment.get(enroll.id)
        else:
            pay = Payment.objects.filter(enrollment=enroll, status=Payment.StatusChoices.COMPLETED).first()
        amt = f"${pay.amount:.2f}" if pay else "$0.00"
        tx_data.append([
            enroll.enrolled_at.strftime('%Y-%m-%d'),
            enroll.student.username[:12],
            enroll.course.title[:25],
            f"{enroll.progress_percent}%",
            amt
        ])

    t_tx = Table(tx_data, colWidths=[80, 90, 200, 70, 80])
    t_tx.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2B6CB0")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ]))
    story.append(t_tx)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()